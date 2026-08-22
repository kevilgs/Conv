import pandas as pd
import os
from config import Config
from .report_formatter import ReportFormatter
from .report_data_processor import ReportDataProcessor
from openpyxl import Workbook

class FinalReportGenerator:
    def __init__(self):
        self.formatter = ReportFormatter()
        self.data_processor = ReportDataProcessor()
    
    def generate_final_report(self, handedover_data, takenover_data, original_filename):
        """Generate final report from processed intermediate data"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Zonal Interchange Report"
            
            # Create report structure
            self.formatter.create_report_structure(ws)
            
            # Get stations in order
            processor = ReportDataProcessor()
            intermediate_file = self._get_intermediate_file_path(original_filename)
            df = pd.read_excel(intermediate_file)
            handedover_stations, takenover_stations = processor.get_stations_in_order(df)
            
            # Get station lists
            handedover_stations = list(handedover_data.keys())
            takenover_stations = list(takenover_data.keys())
            
            BASE_STATION_ORDER = [
                'BSR', 'JL', 'KNW', 'SHRN', 'NAD', 'MKC', 'MTA', 'CNA', 'BEC', 
                'AII', 'HMT', 'BLDI', 'PNU', 'BHU', 'CECC', 'GGM', 'MSH', 'SAUN', 
                'MID_HEADERS', 
                'SAUS', 'MPR', 'GTX', 'PAO', 'NOL', 'BHET', 'SAH', 'SJN',
                'SUBTOTAL'
            ]
            
            all_found_stations = set(handedover_stations + takenover_stations)
            unified_stations = []
            for station in BASE_STATION_ORDER:
                unified_stations.append(station)
                
            for station in sorted(list(all_found_stations)):
                if station not in unified_stations and station not in ['SUBTOTAL', 'GRAND TOTAL', 'MID_HEADERS']:
                    unified_stations.append(station)
            
            # Populate data starting from row 5
            last_data_row = self._populate_report_data(ws, handedover_data, takenover_data, unified_stations, processor)
            
            # Calculate and add totals ONLY at the end
            totals_handed = processor.calculate_totals(handedover_data, unified_stations, is_handedover=True)
            totals_taken = processor.calculate_totals(takenover_data, unified_stations, is_handedover=False)
            
            # Add GRAND TOTAL row
            grand_total_row = last_data_row + 1
            self._add_total_row(ws, grand_total_row, totals_handed, totals_taken, "GRAND TOTAL")
            
            # Create STOCK table after the main report
            ph_stations = self.data_processor._load_ph_stations()
            stock_table_end_row = self.formatter.create_stock_table(ws, grand_total_row, df, ph_stations)
            
            # Calculate custom data and render custom breakdown table
            custom_data = processor.calculate_custom_table_data(df)
            custom_table_end_row = self.formatter.create_custom_breakdown_table(ws, stock_table_end_row, custom_data)
            
            if 'PH_AGGREGATE' in custom_data and custom_data['PH_AGGREGATE']:
                self.formatter.create_ph_aggregate_row(ws, custom_table_end_row + 1, custom_data['PH_AGGREGATE'])
            
            # Save file
            output_filename = self._generate_output_filename(original_filename)
            wb.save(output_filename)
            
            return output_filename, f"Final report generated successfully: {output_filename}"
        
        except Exception as e:
            return None, f"Error generating final report: {str(e)}"
    
    def _populate_report_data(self, ws, handedover_data, takenover_data, unified_stations, processor):
        """Populate the report with actual data preserving order"""
        current_row = 5  # Start from row 5 (after headers)
        
        for station in unified_stations:
            if station == 'MID_HEADERS':
                self.formatter.insert_mid_table_headers(ws, current_row)
                current_row += 3
                continue
                
            if station == 'SUBTOTAL':
                idx = unified_stations.index('MID_HEADERS')
                subtotal_stations = unified_stations[:idx]
                
                totals_handed = processor.calculate_totals(handedover_data, subtotal_stations, is_handedover=True)
                totals_taken = processor.calculate_totals(takenover_data, subtotal_stations, is_handedover=False)
                self._add_total_row(ws, current_row, totals_handed, totals_taken, "SUBTOTAL")
                current_row += 1
                continue
                
            # Calculate how many rows this station group will need
            max_details_length = 1  # At least 1 row for the main station
            
            # Check handedover details
            if station in handedover_data:
                details = handedover_data[station]['details']
                for classification in ['JUMBO', 'BOXN', 'BTPN', 'BTPG', 'SHRA', 'CONT', 'OTHERS', 'EMPTIES']:
                    max_details_length = max(max_details_length, len(details[classification]))
            
            # Check takenover details
            if station in takenover_data:
                details = takenover_data[station]['details']
                for classification in ['JUMBO', 'BOXN', 'BTPN', 'BTPG', 'SHRA', 'CONT', 'OTHERS', 'EMPTIES']:
                    max_details_length = max(max_details_length, len(details[classification]))
                    
            # Dummy details for missing data
            dummy_details = {
                'NO_OF_TRAINS': 0, 'DIESEL': 0, 'JUMBO_LE': '0+0', 'BOXN_LE': '0+0', 'BTPN_LE': '0+0',
                'BOXN_PH_OTH': '0+0', 'CONT_COUNT': '0+0',
                'JUMBO': [], 'BOXN': [], 'BTPN': [], 'BTPG': [], 'SHRA': [], 'CONT': [], 'OTHERS': [], 'EMPTIES': []
            }
            
            # Populate HANDEDOVER section data
            handed_info = handedover_data.get(station, {'ic_sttn': station, 'details': dummy_details})
            
            # Main station info in first row
            ws[f'A{current_row}'] = handed_info['ic_sttn']

            # L+E counts
            details = handed_info['details']
            ws[f'B{current_row}'] = details['NO_OF_TRAINS']
            ws[f'C{current_row}'] = details['DIESEL']
            ws[f'D{current_row}'] = details['JUMBO_LE']
            ws[f'E{current_row}'] = details['BOXN_LE'] 
            ws[f'F{current_row}'] = details['BTPN_LE'] 
            ws[f'G{current_row}'] = details['CONT_COUNT']
            
            # Details section - put each station in separate rows vertically
            self._populate_classification_vertically(ws, current_row, 'H', details['JUMBO'])
            self._populate_classification_vertically(ws, current_row, 'I', details['BOXN'])
            self._populate_classification_vertically(ws, current_row, 'J', details['BTPN'])
            self._populate_classification_vertically(ws, current_row, 'K', details['BTPG'])
            self._populate_classification_vertically(ws, current_row, 'M', details['SHRA'])
            self._populate_classification_vertically(ws, current_row, 'L', details['CONT'])
            self._populate_classification_vertically(ws, current_row, 'N', details['OTHERS'])
            self._populate_classification_vertically(ws, current_row, 'O', details['EMPTIES'])
            
            # Populate TAKENOVER section data
            taken_info = takenover_data.get(station, {'ic_sttn': station, 'details': dummy_details})
            
            # Main station info in first row
            ws[f'P{current_row}'] = taken_info['ic_sttn']

            # L+E counts
            details = taken_info['details']
            ws[f'Q{current_row}'] = details['NO_OF_TRAINS']
            ws[f'R{current_row}'] = details['DIESEL']
            ws[f'S{current_row}'] = details['JUMBO_LE']
            ws[f'T{current_row}'] = details.get('BOXN_PH_OTH', '0+0')
            ws[f'U{current_row}'] = details['BTPN_LE']
            ws[f'V{current_row}'] = details['CONT_COUNT'] 
            
            # Details section - put each station in separate rows vertically
            self._populate_classification_vertically(ws, current_row, 'W', details['JUMBO'])
            self._populate_classification_vertically(ws, current_row, 'X', details['BOXN'])
            self._populate_classification_vertically(ws, current_row, 'Y', details['BTPN'])
            self._populate_classification_vertically(ws, current_row, 'Z', details['BTPG'])
            self._populate_classification_vertically(ws, current_row, 'AB', details['SHRA'])
            self._populate_classification_vertically(ws, current_row, 'AA', details['CONT'])
            self._populate_classification_vertically(ws, current_row, 'AC', details['OTHERS'])
            self._populate_classification_vertically(ws, current_row, 'AD', details['EMPTIES'])
            
            # Apply formatting to all rows used by this station group
            station_start_row = current_row
            station_end_row = current_row + max_details_length - 1

            for row_offset in range(max_details_length):
                self._format_data_row(ws, current_row + row_offset)

            # Apply station-specific formatting with borders and merging
            self.formatter.apply_station_group_formatting(ws, station_start_row, station_end_row, True, True)

            # Apply special formatting to IC STTN and NO OF TRAINS cells
            self.formatter.format_station_and_trains_cells(ws, current_row, True, True)

            # Move to next station group
            current_row += max_details_length


        # Return the last row used
        return current_row - 1  # Return last data row
    
    def _populate_classification_vertically(self, ws, start_row, column, station_list):
        """Populate classification details in separate cells vertically"""
        if not station_list:
            return
        
        for idx, station in enumerate(station_list):
            cell_row = start_row + idx
            ws[f'{column}{cell_row}'] = station
            
            # Apply formatting to the detail cell
            cell = ws[f'{column}{cell_row}']
            cell.font = self.formatter.normal_font
            cell.alignment = self.formatter.center_align
            cell.border = self.formatter.thin_border
            cell.fill = self.formatter.white_fill

    def _format_data_row(self, ws, row_num):
        """Apply formatting to a data row"""
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O',
                   'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
            cell = ws[f'{col}{row_num}']
            cell.font = self.formatter.normal_font
            cell.alignment = self.formatter.center_align
            cell.border = self.formatter.thin_border
            cell.fill = self.formatter.white_fill
    
    def _add_total_row(self, ws, current_row, totals_handed, totals_taken, label):
        """Add a total row with calculated totals"""
        # HANDEDOVER totals
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = f"{totals_handed['NO_OF_TRAINS'][0]}/{totals_handed['NO_OF_TRAINS'][1]}"
        ws[f'C{current_row}'] = totals_handed['DIESEL']
        ws[f'D{current_row}'] = f"{totals_handed['JUMBO_LE'][0]}+{totals_handed['JUMBO_LE'][1]}"
        ws[f'E{current_row}'] = f"{totals_handed['BOXN_LE'][0]}+{totals_handed['BOXN_LE'][1]}"
        ws[f'F{current_row}'] = f"{totals_handed['BTPN_LE'][0]}+{totals_handed['BTPN_LE'][1]}"
        ws[f'G{current_row}'] = totals_handed['CONT']
        
        # TAKENOVER totals
        ws[f'P{current_row}'] = label
        ws[f'Q{current_row}'] = f"{totals_taken['NO_OF_TRAINS'][0]}/{totals_taken['NO_OF_TRAINS'][1]}"
        ws[f'R{current_row}'] = totals_taken['DIESEL']
        ws[f'S{current_row}'] = f"{totals_taken['JUMBO_LE'][0]}+{totals_taken['JUMBO_LE'][1]}"
        ws[f'T{current_row}'] = f"{totals_taken['BOXN_PH_OTH'][0]}+{totals_taken['BOXN_PH_OTH'][1]}"
        ws[f'U{current_row}'] = f"{totals_taken['BTPN_LE'][0]}+{totals_taken['BTPN_LE'][1]}"
        ws[f'V{current_row}'] = totals_taken['CONT']
        
        # Apply bold formatting and borders to ALL cells in the row
        from openpyxl.styles import Font
        bold_font = Font(name='Arial', size=9, bold=True)
        
        # Format all columns A through AD
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O',
                   'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
            cell = ws[f'{col}{current_row}']
            cell.font = bold_font
            cell.alignment = self.formatter.center_align
            cell.border = self.formatter.thin_border
            cell.fill = self.formatter.white_fill
        
        self.formatter.format_station_and_trains_cells(ws, current_row, True, True)  # Add this line
        
        # Apply thick border around the entire total row
        self.formatter.apply_thick_border_to_row(ws, current_row, 'A', 'AD')
    
    def _get_intermediate_file_path(self, original_filename):
        """Get the path to the intermediate XLSX file"""
        from config import Config
        import os
        
        if original_filename.endswith('_processed.xlsx'):
            return os.path.join(Config.INTERMEDIATE_FOLDER, original_filename)
            
        base_name = os.path.splitext(original_filename)[0]
        intermediate_filename = f"{base_name}_processed.xlsx"
        return os.path.join(Config.INTERMEDIATE_FOLDER, intermediate_filename)

    def _generate_output_filename(self, original_filename):
        """Generate output filename for final report"""
        from config import Config
        import os
        
        base_name = os.path.splitext(original_filename)[0]
        if base_name.endswith('_processed'):
            base_name = base_name[:-10] # remove '_processed'
            
        output_filename = f"{base_name}_final_report.xlsx"
        
        # Create reports directory if it doesn't exist
        os.makedirs(Config.REPORTS_FOLDER, exist_ok=True)
        
        return os.path.join(Config.REPORTS_FOLDER, output_filename)