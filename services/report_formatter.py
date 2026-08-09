import pandas as pd
from datetime import datetime,timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReportFormatter:
    def __init__(self):
        # Font styles
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.section_header_font = Font(name='Arial', size=12, bold=True)
        self.column_header_font = Font(name='Arial', size=10, bold=True)
        self.normal_font = Font(name='Arial', size=9)
        
        # No fill colors - all white
        self.white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Alignment
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.vertical_align = Alignment(horizontal='center', vertical='center', text_rotation=90)  # Add this line
        
        # Borders
        self.thick_border = Border(
            left=Side(style='thick'), right=Side(style='thick'),
            top=Side(style='thick'), bottom=Side(style='thick')
        )
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def create_report_structure(self, ws):
        """Create the complete report structure with headers and formatting"""
        
        # 1st Row - Main Title
        current_date = (datetime.now() - timedelta(days=1)).strftime('%d-%m-%Y')
        ws['A1'] = f"ZONAL INTERCHANGE ON {current_date}"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_align
        ws['A1'].fill = self.white_fill
        ws.merge_cells('A1:AD1')
        
        # 2nd Row - Section Headers
        ws['A2'] = "HANDEDOVER"
        ws['A2'].font = self.section_header_font
        ws['A2'].alignment = self.center_align
        ws['A2'].fill = self.white_fill
        ws.merge_cells('A2:O2')
        
        ws['P2'] = "TAKENOVER"
        ws['P2'].font = self.section_header_font
        ws['P2'].alignment = self.center_align
        ws['P2'].fill = self.white_fill
        ws.merge_cells('P2:AD2')
        
        # 3rd Row - Main Headers
        self._create_row3_headers(ws)
        
        # 4th Row - Sub Headers
        self._create_row4_headers(ws)
        
        # Set column widths and row heights
        self._set_dimensions(ws)
        
        # Apply borders to all header cells
        self._apply_header_borders(ws)
    
    def _create_row3_headers(self, ws):
        """Create 3rd row headers"""
        # HANDEDOVER section headers
        handedover_headers = {
            'A3': 'IC STTN',
            'B3': 'NO OF TRAINS',
            'C3': 'DIESEL',
            'D3': 'JUMBO',
            'E3': 'BOXN',
            'F3': 'BTPN',
            'G3': 'CONT',
            'H3': 'DETAILS'
        }
        
        for cell, title in handedover_headers.items():
            ws[cell] = title
            ws[cell].font = self.column_header_font
            ws[cell].fill = self.white_fill
            
            # Apply vertical alignment to all except DETAILS
            if title == 'DETAILS':
                ws[cell].alignment = self.center_align  # Keep horizontal
            else:
                ws[cell].alignment = self.vertical_align  # Make vertical
        
        # TAKENOVER section headers
        takenover_headers = {
            'P3': 'IC STTN',
            'Q3': 'NO OF TRAINS',
            'R3': 'DIESEL',
            'S3': 'JUMBO',
            'T3': 'BOXN',
            'U3': 'BTPN',
            'V3': 'CONT',
            'W3': 'DETAILS'
        }
        
        for cell, title in takenover_headers.items():
            ws[cell] = title
            ws[cell].font = self.column_header_font
            ws[cell].fill = self.white_fill
            
            # Apply vertical alignment to all except DETAILS
            if title == 'DETAILS':
                ws[cell].alignment = self.center_align  # Keep horizontal
            else:
                ws[cell].alignment = self.vertical_align  # Make vertical
        
        # Merge cells for row 3
        self._merge_row3_cells(ws)
    
    def _merge_row3_cells(self, ws):
        """Merge cells for row 3"""
        # HANDEDOVER merges
        ws.merge_cells('A3:A4')  # IC STTN
        ws.merge_cells('B3:B4')  # NO OF TRAINS
        ws.merge_cells('C3:C4')  # DIESEL
        ws.merge_cells('G3:G4')  # CONT
        ws.merge_cells('H3:O3')  # DETAILS
        
        # TAKENOVER merges
        ws.merge_cells('P3:P4')  # IC STTN
        ws.merge_cells('Q3:Q4')  # NO OF TRAINS
        ws.merge_cells('R3:R4')  # DIESEL
        ws.merge_cells('V3:V4')  # CONT
        ws.merge_cells('W3:AD3')  # DETAILS
    
    def _create_row4_headers(self, ws):
        """Create 4th row sub-headers"""
        # HANDEDOVER sub-headers
        handedover_row4 = {
            'D4': 'L+E',
            'E4': 'L+E',
            'F4': 'L+E',
            'H4': 'JUMBO',
            'I4': 'BOXN',
            'J4': 'BTPN',
            'K4': 'BTPG',
            'L4': 'CONT',
            'M4': 'SHRA',
            'N4': 'OTHERS',
            'O4': 'EMPTIES'
        }
        
        for cell, title in handedover_row4.items():
            ws[cell] = title
            ws[cell].font = self.column_header_font
            ws[cell].alignment = self.vertical_align  # All vertical in row 4
            ws[cell].fill = self.white_fill
        
        # TAKENOVER sub-headers
        takenover_row4 = {
            'S4': 'L+E',
            'T4': 'PH+OTH',
            'U4': 'L+E',
            'W4': 'JUMBO',
            'X4': 'BOXN',
            'Y4': 'BTPN',
            'Z4': 'BTPG',
            'AA4': 'CONT',
            'AB4': 'SHRA',
            'AC4': 'OTHERS',
            'AD4': 'EMPTIES'
        }
        
        for cell, title in takenover_row4.items():
            ws[cell] = title
            ws[cell].font = self.column_header_font
            ws[cell].alignment = self.vertical_align  # All vertical in row 4
            ws[cell].fill = self.white_fill
    
    def _set_dimensions(self, ws):
        """Set column widths and row heights"""
        # Column widths
        column_widths = {
            'A': 12, 'B': 12, 'C': 10, 'D': 8, 'E': 8, 'F': 8, 'G': 8,
            'H': 10, 'I': 8, 'J': 8, 'K': 8, 'L': 8, 'M': 8, 'N': 10, 'O': 10,
            'P': 12, 'Q': 12, 'R': 10, 'S': 8, 'T': 10, 'U': 8, 'V': 8,
            'W': 10, 'X': 8, 'Y': 8, 'Z': 8, 'AA': 8, 'AB': 8, 'AC': 10, 'AD': 10
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 20
    
    def _apply_header_borders(self, ws):
        """Apply borders to all header cells"""
        # Apply borders to all cells from A1 to AD4
        for row in range(1, 5):
            for col in range(1, 31):  # A to AD is 30 columns
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                # Ensure all cells have white fill
                cell.fill = self.white_fill
    
    def apply_station_group_formatting(self, ws, start_row, end_row, has_handedover=True, has_takenover=True):
        """Apply thick border around station group and merge single-value columns"""
        from openpyxl.styles import Border, Side
        
        thick_side = Side(style='thick')
        
        # Apply thick border around entire station group
        self._apply_thick_border_to_range(ws, start_row, end_row, 'A', 'AD')
        
        # Merge single-value columns for handedover section
        if has_handedover and end_row > start_row:
            merge_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']  # IC STTN through CONT
            for col in merge_columns:
                ws.merge_cells(f'{col}{start_row}:{col}{end_row}')
                # Center the merged cell content
                merged_cell = ws[f'{col}{start_row}']
                merged_cell.alignment = self.center_align
        
        # Merge single-value columns for takenover section  
        if has_takenover and end_row > start_row:
            merge_columns = ['P', 'Q', 'R', 'S', 'T', 'U', 'V']  # IC STTN through CONT
            for col in merge_columns:
                ws.merge_cells(f'{col}{start_row}:{col}{end_row}')
                # Center the merged cell content
                merged_cell = ws[f'{col}{start_row}']
                merged_cell.alignment = self.center_align

    def _apply_thick_border_to_range(self, ws, start_row, end_row, start_col, end_col):
        """Apply thick border around a range of cells"""
        from openpyxl.styles import Border, Side
        from openpyxl.utils import column_index_from_string
        
        thick_side = Side(style='thick')
        
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                cell = ws.cell(row=row, column=col_idx)
                
                # Determine which borders should be thick
                left_thick = (col_idx == start_col_idx)
                right_thick = (col_idx == end_col_idx)
                top_thick = (row == start_row)
                bottom_thick = (row == end_row)
                
                cell.border = Border(
                    left=thick_side if left_thick else cell.border.left,
                    right=thick_side if right_thick else cell.border.right,
                    top=thick_side if top_thick else cell.border.top,
                    bottom=thick_side if bottom_thick else cell.border.bottom
                )
    
    def apply_thick_border_to_row(self, ws, row, start_col, end_col):
        """Apply thick border around an entire row"""
        from openpyxl.styles import Border, Side
        from openpyxl.utils import column_index_from_string
        
        thick_side = Side(style='thick')
        
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        for col_idx in range(start_col_idx, end_col_idx + 1):
            cell = ws.cell(row=row, column=col_idx)
            
            # Apply thick border on all sides
            cell.border = Border(
                left=thick_side if col_idx == start_col_idx else cell.border.left,
                right=thick_side if col_idx == end_col_idx else cell.border.right,
                top=thick_side,
                bottom=thick_side
            )
    
    def format_station_and_trains_cells(self, ws, row, has_handedover, has_takenover):
        """Apply special formatting to IC STTN and NO OF TRAINS cells - size 14, bold"""
        from openpyxl.styles import Font
        
        # Create font: size 14, bold
        station_trains_font = Font(name='Arial', size=14, bold=True)
        
        if has_handedover:
            # Format handedover IC STTN (column A) and NO OF TRAINS (column B)
            ws[f'A{row}'].font = station_trains_font
            ws[f'B{row}'].font = station_trains_font
        
        if has_takenover:
            # Format takenover IC STTN (column P) and NO OF TRAINS (column Q)
            ws[f'P{row}'].font = station_trains_font
            ws[f'Q{row}'].font = station_trains_font
    
    def create_stock_table(self, ws, start_row, df, ph_stations):
        """Create STOCK table with OB, H/O, T/O, CB columns auto-filled"""
        # Leave some gap (3 rows) from last data
        table_start_row = start_row + 3
        
        # Create table headers
        headers = ['STOCK', 'OB', 'H/O', 'T/O', 'CB']
        header_columns = ['A', 'B', 'C', 'D', 'E']
        
        # Set headers
        for i, (col, header) in enumerate(zip(header_columns, headers)):
            cell = ws[f'{col}{table_start_row}']
            cell.value = header
            cell.font = self.column_header_font
            cell.alignment = self.center_align
            cell.fill = self.white_fill
            cell.border = self.thin_border
        
        # Create stock rows
        stock_items = ['JUMBO', 'BOXN', 'BTPN', 'CONT', 'SHRA']
        
        for i, stock_item in enumerate(stock_items):
            row = table_start_row + 1 + i
            
            # Stock name in column A
            ws[f'A{row}'].value = stock_item
            ws[f'A{row}'].font = self.normal_font
            ws[f'A{row}'].alignment = self.center_align
            ws[f'A{row}'].fill = self.white_fill
            ws[f'A{row}'].border = self.thin_border
            
            # Calculate H/O and T/O values
            search_class = 'BOX' if stock_item == 'BOXN' else stock_item
            
            ho_df = df[df['HANDEDOVER CLASSIFICATION'] == search_class]
            ho_total = len(ho_df)
            ho_empty = len(ho_df[ho_df['HANDED OVER L/E'] == 'E'])
            
            to_df = df[df['TAKENOVER CLASSIFICATION'] == search_class]
            to_total = len(to_df)
            to_empty = len(to_df[to_df['TAKEN OVER L/E'] == 'E'])
            
            if stock_item == 'BOXN':
                to_l_df = to_df[to_df['TAKEN OVER L/E'] == 'L']
                to_empty = len(to_l_df[to_l_df['TAKEN OVER STTN TO'].isin(ph_stations)])
                
            ho_val = f"{ho_total}/{ho_empty}"
            to_val = f"{to_total}/{to_empty}"
            
            # Cells for OB, H/O, T/O, CB (columns B, C, D, E)
            for col in ['B', 'C', 'D', 'E']:
                cell = ws[f'{col}{row}']
                if col == 'B':
                    cell.value = ""  # Empty for manual entry
                elif col == 'C':
                    cell.value = ho_val
                elif col == 'D':
                    cell.value = to_val
                elif col == 'E':
                    cell.value = f'=IFERROR(VALUE(B{row}), 0) + IFERROR(VALUE(LEFT(D{row}, FIND("/", D{row})-1)), IFERROR(VALUE(D{row}), 0)) - IFERROR(VALUE(LEFT(C{row}, FIND("/", C{row})-1)), IFERROR(VALUE(C{row}), 0))'
                
                cell.font = self.normal_font
                cell.alignment = self.center_align
                cell.fill = self.white_fill
                cell.border = self.thin_border
        
        # Set column widths for stock table
        stock_column_widths = {'A': 15, 'B': 12, 'C': 12, 'D': 12, 'E': 12}
        for col, width in stock_column_widths.items():
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width, width)
        
        # Return the last row of the table for future reference
        return table_start_row + len(stock_items)

    def create_custom_breakdown_table(self, ws, start_row, custom_data):
        """Create custom breakdown table for JUMBO, BOXN, BTPN"""
        current_row = start_row + 2
        
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        from openpyxl.cell.text import InlineFont

        # Bold fonts
        bold_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=9)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        bold_inline = InlineFont(b=True, sz=9, rFont='Arial')
        normal_inline = InlineFont(b=False, sz=9, rFont='Arial')

        def format_cell(cell, value, font, align, border):
            cell.value = value
            if not isinstance(value, CellRichText):
                cell.font = font
            cell.alignment = align
            cell.border = border
            cell.fill = self.white_fill

        def apply_border_to_range(ws, min_row, min_col, max_row, max_col, border):
            """Helper to apply border to all cells in a merged range and set background"""
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    cell.fill = self.white_fill

        def write_stations(ws, row, col_start_idx, stations_list):
            """Write stations as a single comma-separated string in a horizontally merged cell, with bold IC STTN"""
            elements = []
            for i, sttn_str in enumerate(stations_list):
                if ' - [' in sttn_str:
                    ic_part, rest = sttn_str.split(' - [', 1)
                    rest = ' - [' + rest
                    elements.append(TextBlock(bold_inline, ic_part))
                    if i < len(stations_list) - 1:
                        rest += ", "
                    elements.append(TextBlock(normal_inline, rest))
                elif '[' in sttn_str:
                    ic_part, rest = sttn_str.split('[', 1)
                    rest = '[' + rest
                    elements.append(TextBlock(bold_inline, ic_part))
                    if i < len(stations_list) - 1:
                        rest += ", "
                    elements.append(TextBlock(normal_inline, rest))
                else:
                    if i < len(stations_list) - 1:
                        sttn_str += ", "
                    elements.append(TextBlock(normal_inline, sttn_str))
                    
            val = CellRichText(*elements) if elements else ""
            end_col = 12  # Merge up to column 12
            
            ws.merge_cells(start_row=row, start_column=col_start_idx, end_row=row, end_column=end_col)
            apply_border_to_range(ws, row, col_start_idx, row, end_col, thin_border)
            
            cell = ws.cell(row=row, column=col_start_idx)
            format_cell(cell, val, normal_font, left_align, thin_border)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for class_name in ['JUMBO', 'BOXN', 'BTPN']:
            data = custom_data[class_name]
            
            # 1. Main Title Row
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
            apply_border_to_range(ws, current_row, 1, current_row, 12, thin_border)
            format_cell(ws.cell(row=current_row, column=1), class_name, bold_font, center_align, thin_border)
            current_row += 1
            
            # 2. Header Row (TOTAL and EMPTY/PH)
            apply_border_to_range(ws, current_row, 1, current_row, 12, thin_border)
            format_cell(ws.cell(row=current_row, column=2), "TOTAL", bold_font, center_align, thin_border)
            format_cell(ws.cell(row=current_row, column=3), "EMPTY", bold_font, center_align, thin_border)
            current_row += 1
            
            # 3. H/O Row
            apply_border_to_range(ws, current_row, 1, current_row, 12, thin_border)
            format_cell(ws.cell(row=current_row, column=1), "H/O", bold_font, center_align, thin_border)
            format_cell(ws.cell(row=current_row, column=2), data['HO']['total'], normal_font, center_align, thin_border)
            format_cell(ws.cell(row=current_row, column=3), data['HO']['empty_total'], normal_font, center_align, thin_border)
            
            format_cell(ws.cell(row=current_row, column=4), "E", bold_font, center_align, thin_border)
            format_cell(ws.cell(row=current_row, column=5), data['HO']['E']['total'], normal_font, center_align, thin_border)
            write_stations(ws, current_row, 6, data['HO']['E']['stations'])
            
            current_row += 1
            
            # 4. T/O Section
            to_start_row = current_row
            num_to_rows = 4 if class_name == 'BOXN' else 3
            to_end_row = to_start_row + num_to_rows - 1
            
            # Initialize borders for the whole block
            apply_border_to_range(ws, to_start_row, 1, to_end_row, 12, thin_border)
            
            # Merge and populate Col 1 (T/O)
            ws.merge_cells(start_row=to_start_row, start_column=1, end_row=to_end_row, end_column=1)
            format_cell(ws.cell(row=to_start_row, column=1), "T/O", bold_font, center_align, thin_border)
            
            # Merge and populate Col 2 (TOTAL)
            ws.merge_cells(start_row=to_start_row, start_column=2, end_row=to_end_row, end_column=2)
            format_cell(ws.cell(row=to_start_row, column=2), data['TO']['total'], normal_font, center_align, thin_border)
            
            # Merge and populate Col 3 (EMPTY/PH)
            ws.merge_cells(start_row=to_start_row, start_column=3, end_row=to_end_row, end_column=3)
            empty_val = data['TO']['ph_total'] if class_name == 'BOXN' else data['TO']['empty_total']
            format_cell(ws.cell(row=to_start_row, column=3), empty_val, normal_font, center_align, thin_border)
            
            # T/O E
            format_cell(ws.cell(row=current_row, column=4), "E", bold_font, center_align, thin_border)
            format_cell(ws.cell(row=current_row, column=5), data['TO']['E']['total'], normal_font, center_align, thin_border)
            write_stations(ws, current_row, 6, data['TO']['E']['stations'])
            current_row += 1
            
            # T/O PU
            format_cell(ws.cell(row=current_row, column=4), "PU", bold_font, center_align, thin_border)
            format_cell(ws.cell(row=current_row, column=5), data['TO']['PU']['total'], normal_font, center_align, thin_border)
            write_stations(ws, current_row, 6, data['TO']['PU']['stations'])
            current_row += 1
            
            # T/O B/P
            format_cell(ws.cell(row=current_row, column=4), "B/P", bold_font, center_align, thin_border)
            format_cell(ws.cell(row=current_row, column=5), data['TO']['B/P']['total'], normal_font, center_align, thin_border)
            write_stations(ws, current_row, 6, data['TO']['B/P']['stations'])
            current_row += 1
            
            # T/O PH (Only for BOXN)
            if class_name == 'BOXN':
                format_cell(ws.cell(row=current_row, column=4), "PH", bold_font, center_align, thin_border)
                format_cell(ws.cell(row=current_row, column=5), data['TO']['PH']['total'], normal_font, center_align, thin_border)
                write_stations(ws, current_row, 6, data['TO']['PH']['stations'])
                current_row += 1
            
            # Add a spacer row between tables
            current_row += 1
            
        return current_row

